import streamlit as st
import pandas as pd
import sys
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border

# ---------------------------
# Streamlit UI
# ---------------------------
st.set_page_config(layout="wide", page_title="Kabaddi QC Tool")

# st.title("Kabaddi Data Processing & QC Tool - Old Dashboard")
st.markdown(
    '<h1>Kabaddi Data Processing & QC tool - <span style="color:yellow;">PKL (LongoMatch)</span></h1>',
    unsafe_allow_html=True)

###################
match_no = st.text_input("Enter Match Number", value = "0")
###################

st.markdown("")

# --- Upload CSV ---
uploaded_file = st.file_uploader("Upload raw Kabaddi CSV, process it, and download the cleaned output (Excel file).", type=["csv"])

if uploaded_file:
    # Store raw_df in session_state for safe access
    st.session_state.raw_df = pd.read_csv(uploaded_file, delimiter=';', header=None, dtype=str, skiprows=1)

    # --- Show Total Rows and Columns ---
    rows, cols = st.session_state.raw_df.shape if st.session_state.raw_df is not None else (0, 0)
    st.write(f"**RAW File: Total rows:** `{rows}` | **Total columns:** `{cols}`")

    # --- Show first 5 rows of raw file ---
    st.subheader("Raw File Preview")
    st.dataframe(st.session_state.raw_df, height=210)

    # CSS to style the Process button
    st.markdown(
    """
    <style>
    div.stButton>button {
        color: yellow !important;
        font-weight: bolder !important;
        font-size: 30px !important;  /* Increase font size */
        background-color: black !important;
        border: none !important;
        padding: 10px 20px !important; /* Makes button bigger */
    }
    </style>
    """,
    unsafe_allow_html=True)

    # --- Process Button ---
    if st.button("Process CSV", use_container_width=True):

        st.subheader("Quality Check Logs")
        log_output = io.StringIO()
        sys.stdout = log_output  # Capture all print statements

        try:
            # Define raw_df before using it
            raw_df = st.session_state.raw_df.copy()

            # Step 2: Find the row where the first column is strictly "Name"
            header_row_idx_search = raw_df[raw_df.iloc[:, 0].astype(str).str.strip() == "Name"].index

            if header_row_idx_search.empty:
                print("❌ Could not find a row strictly equal to 'Name'.")
                sys.exit()

            header_row_idx = header_row_idx_search[0]

            # Step 3: Use that row as the header, and keep only the rows below it
            df = raw_df.copy()
            df.columns = df.iloc[header_row_idx].astype(str).str.strip()
            df = df.iloc[header_row_idx + 1:].reset_index(drop=True)

            # Step 4: Keep only rows where first column strictly starts with "Raid "
            df = df[df.iloc[:, 0].astype(str).str.strip().str.startswith("Raid ")].reset_index(drop=True)

            if df.empty:
                print("❌ No rows found strictly starting with 'Raid '.")
                sys.exit()

             # Step 5: Rename Columns
            new_col_names = [
                'Name','Time','Start','Stop','Team','Player','Raid 1','Raid 2','Raid 3',
                'D1','D2','D3','D4','D5','D6','D7','Successful','Empty','Unsuccessful',
                'Bonus','No Bonus','Z1','Z2','Z3','Z4','Z5','Z6','Z7','Z8','Z9','RT0',
                'RT1','RT2','RT3','RT4','RT5','RT6','RT7','RT8','RT9','DT0','DT1','DT2',
                'DT3','DT4','Hand touch','Running hand touch','Toe touch','Running Kick',
                'Reverse Kick','Side Kick','Defender self out','Body hold',
                'Ankle hold','Single Thigh hold','Push','Dive','DS0','DS1','DS2','DS3','In Turn',
                'Out Turn','Create Gap','Jump','Dubki','Struggle','Release','Block','Chain_def','Follow',
                'Technical Point Raiding','All Out', *(f'RL{i}' for i in range(1, 31)),
                'Raider self out','Running Bonus','Centre Bonus','LCorner','LIN','LCover','Center',
                'RCover','RIN','RCorner','Flying Touch','Double Thigh Hold','Flying Reach','Clean','Not Clean',
                'Yes','No','Z10','Z11','First Half','Second Half','Technical Point Defending']

            if len(df.columns) == len(new_col_names):
                df.columns = new_col_names
            else:
                print(f"❌ Column mismatch: got {len(df.columns)}, expected {len(new_col_names)}")
                sys.exit()

            # =========================================================================
            # START: Part 2 - Transformation and QCs
            # =========================================================================
            
           # ---------------- Drop unused columns ----------------
            df.drop(['Time', 'Team'], axis=1, inplace=True, errors='ignore')


            # -------- Raid_Number --------
    
            for c in ['Raid 1', 'Raid 2', 'Raid 3']:
                df[c] = pd.to_numeric(df[c].astype(str).str.strip().replace('', '0'), errors='coerce').fillna(0).astype(int)

            df.loc[df['Raid 2'] == 1, 'Raid 2'] = 2
            df.loc[df['Raid 3'] == 1, 'Raid 3'] = 3

            df['Raid_Number'] = df['Raid 1'] + df['Raid 2'] + df['Raid 3']
            df.drop(['Raid 1', 'Raid 2', 'Raid 3'], axis=1, inplace=True)


            # ------ Rename key columns ------
        
            df.rename(columns={'Name': 'Event_Number',
                               'Technical Point Raiding': 'Technical_Point_Raiding_Team',
                               'Technical Point Defending': 'Technical_Point_Defending_Team',
                               'All Out': 'All_Out'}, inplace=True)

            # ------ Number_of_Defenders ------

            defender_cols = ['D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7']

            for idx, col in enumerate(defender_cols, 1):

                df[col] = pd.to_numeric(df[col].astype(str).str.strip().replace('', '0'),
                                        errors='coerce').fillna(0).astype(int)
                
                df[col] = (df[col] == 1).astype(int) * idx

            df['Number_of_Defenders'] = df[defender_cols].sum(axis=1).astype(int)
            df.drop(columns=defender_cols, inplace=True)


            # ------ Outcome ------

            # 1. Ensure numeric conversion
            for col in ['Successful', 'Empty', 'Unsuccessful']:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → label, 0 → empty string
            df['Successful'] = df['Successful'].map({1: 'Successful', 0: ''})
            df['Empty'] = df['Empty'].map({1: 'Empty', 0: ''})
            df['Unsuccessful'] = df['Unsuccessful'].map({1: 'Unsuccessful', 0: ''})

            # 3. Safely join non-empty labels
            df['Outcome'] = df[['Successful', 'Empty', 'Unsuccessful']].apply(lambda row: ' '.join(filter(None, row)), axis=1)

            df.drop(['Successful', 'Empty', 'Unsuccessful'], axis=1, inplace=True)


            # ------ Bonus ------

            df_bonus = df[['Bonus', 'No Bonus', 'Centre Bonus', 'Running Bonus']].copy()

            # Convert to integers to avoid string concatenation issues
            for col in df_bonus.columns:
                df_bonus[col] = pd.to_numeric(df_bonus[col], errors='coerce').fillna(0).astype(int)

            # Create unified "Bonus" indicator
            df_bonus['Bonus'] = df_bonus[['Bonus', 'Centre Bonus', 'Running Bonus']].max(axis=1)
            df_bonus['Bonus'] = df_bonus['Bonus'].map({1: 'Yes', 0: ''})
            df_bonus['No Bonus'] = df_bonus['No Bonus'].map({1: 'No', 0: ''})

            # Combine cleanly
            df_bonus['Bonus'] = (df_bonus['Bonus'] + ' ' + df_bonus['No Bonus']).str.strip()

            # If all are 0 → set Bonus to "No"
            df_bonus.loc[(df_bonus[['Bonus', 'No Bonus']] == '').all(axis=1),'Bonus'] = 'No'

            df_bonus.drop(columns=['No Bonus', 'Centre Bonus', 'Running Bonus'], inplace=True)


            # ------ Type_of_Bonus ------

            bonus_cols = ['Bonus', 'Centre Bonus', 'Running Bonus']

            # Ensure numeric 0/1
            for col in bonus_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # Map 1 → column name, 0 → blank
            for col in bonus_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # Join them safely
            df['Type_of_Bonus'] = df[bonus_cols].apply(lambda x: ' '.join(filter(None, x)), axis=1)

            # Drop original raw bonus columns
            df.drop(columns=bonus_cols + ['No Bonus'], inplace=True, errors='ignore')

            # Merge final clean Bonus column back into main df
            df = pd.concat([df_bonus, df], axis=1)


            # ------ Zone_of_Action ------

            zone_cols = ['Z1', 'Z2', 'Z3', 'Z4', 'Z5', 'Z6', 'Z7', 'Z8', 'Z9', 'Z10', 'Z11']

            # Convert to integers first (handles '0', '1', blanks)
            for col in zone_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # Replace 1 → column name, 0 → blank
            for col in zone_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # Join zone names cleanly
            df['Zone_of_Action'] = df[zone_cols].apply(lambda x: ' '.join(filter(None, x)), axis=1)

            df.drop(columns=zone_cols, inplace=True)


            # ------ Raiding_Team_Points ------

            rt_cols = ['RT0', 'RT1', 'RT2', 'RT3', 'RT4', 'RT5', 'RT6', 'RT7', 'RT8', 'RT9']

            # Convert to integers first
            for col in rt_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # Replace 1 → its numeric suffix (e.g., RT3 → 3)
            for col in rt_cols:
                num = int(col.replace("RT", ""))
                df[col] = df[col].map({1: num, 0: 0})

            # Sum up points
            df['Raiding_Team_Points'] = df[rt_cols].sum(axis=1).astype(int)
            df.drop(columns=rt_cols, inplace=True)


            # ----------- Defending_Team_Points -----------

            dt_cols = ['DT0', 'DT1', 'DT2', 'DT3', 'DT4']

            for col in dt_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                num = int(col.replace("DT", ""))
                df[col] = df[col].map({1: num, 0: 0})

            df['Defending_Team_Points'] = df[dt_cols].sum(axis=1).astype(int)
            df.drop(columns=dt_cols, inplace=True)


            # ------ Attacking_Skill ----------

            att_skill_cols = ['Hand touch', 'Running hand touch', 'Toe touch', 'Running Kick', 'Reverse Kick',
                                'Side Kick', 'Defender self out', 'Flying Touch']

            # 1. Clean and convert to integers (0/1)
            for col in att_skill_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in att_skill_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['Attacking_Skill'] = df[att_skill_cols].apply(lambda x: ', '.join(filter(None, x)), axis=1)

            df.drop(columns=att_skill_cols, inplace=True)

            
            # ------------- Defensive_Skill --------------

            ds_skill_cols = ['Body hold', 'Ankle hold', 'Single Thigh hold', 'Double Thigh Hold', 'Push', 'Dive', 'Block',
                                'Chain_def', 'Follow', 'Raider self out']

            # 1. Clean and convert to integers (0/1)
            for col in ds_skill_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in ds_skill_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['Defensive_Skill'] = df[ds_skill_cols].apply(lambda x: ', '.join(filter(None, x)), axis=1)

            df.drop(columns=ds_skill_cols, inplace=True)


            # ------------ Number_of_Defenders_Self_Out --------------

            dso_cols = ['DS0', 'DS1', 'DS2', 'DS3']

            for col in dso_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                num = int(col.replace("DS", ""))
                df[col] = df[col].map({1: num, 0: 0})

            df['Number_of_Defenders_Self_Out'] = df[dso_cols].sum(axis=1).astype(int)
            df.drop(columns=dso_cols, inplace=True)

            
            # ------ Counter_Action_Skill ------

            ca_cols = ['In Turn', 'Out Turn', 'Create Gap', 'Jump', 'Dubki', 'Struggle', 'Release', 'Flying Reach']

            # 1. Clean and convert to integers (0/1)
            for col in ca_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in ca_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['Counter_Action_Skill'] = df[ca_cols].apply(lambda x: ', '.join(filter(None, x)), axis=1)

            df.drop(columns=ca_cols, inplace=True)


            # ------ Defender_Positions ------

            def_pos_cols = ['LCorner', 'LIN', 'LCover', 'Center', 'RCover', 'RIN', 'RCorner']

            # 1. Clean and convert to integers (0/1)
            for col in def_pos_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in def_pos_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['Defender_Position'] = (df[def_pos_cols].apply(lambda row: ', '.join(filter(None, row)), axis=1))
            
            df.drop(columns=def_pos_cols, inplace=True)

            
            # ------ QoD_Skill ------

            qod_cols = ['Clean', 'Not Clean']
            # 1. Clean and convert to integers (0/1)
            for col in qod_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in qod_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['QoD_Skill'] = df[qod_cols].apply(lambda x: ', '.join(filter(None, x)), axis=1)

            df.drop(columns=qod_cols, inplace=True)


            # ---------------- Raiding Length ----------------

            rl_cols = [f'RL{i}' for i in range(1, 31)]

            for col in rl_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                num = int(col.replace("RL", ""))
                df[col] = df[col].map({1: num, 0: 0})

            # Calculate Actual Raid_Length
            df['Raid_Length'] = 30 - df[rl_cols].sum(axis=1).astype(int)
            df.drop(columns=rl_cols, inplace=True)

                        
            # ------------ Half ------------

            half_cols = ['First Half', 'Second Half']

            # 1. Clean and convert to integers (0/1)
            for col in half_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in half_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['Half'] = df[half_cols].apply(lambda x: ', '.join(filter(None, x)), axis=1)

            # 4 Remove ' Half'
            df['Half'] = df['Half'].str.replace(' Half', '', regex=False)

            df.drop(columns = half_cols, inplace=True)


            # ------- Define IDs --------

            n = len(df)
            df['Tournament_ID'] = "T001"
            df['Season_ID'] = "S13"
            df['Match_No'] = int(match_no)
            df['Match_ID'] = f"M{int(match_no):03d}"
            df['Match_Raid_Number'] = range(1, n + 1)

  
            # ---------------- Raider & Defenders Names ----------------

            # Split player column
            names = (df['Player'].str.split(r'\s*\|\s*', expand=True)
                    .apply(lambda s: s.str.split('-', n=1).str[1].str.strip().str.title()))

            # Ensure exactly 8 columns (1 raider + 7 defenders)
            names = names.reindex(columns=range(8))

            # Rename columns
            names.columns = ['Raider_Name'] + [f'Defender_{i}_Name' for i in range(1, 8)]

            # Replace original column
            df = df.drop(columns='Player').join(names)
            
        
            # ---------------- Time -----------------

            # Remove milliseconds
            start_str = df['Start'].str.split(',').str[0]
            stop_str  = df['Stop'].str.split(',').str[0]

            # Convert mm:ss → hh:mm:ss
            start = pd.to_timedelta('00:' + start_str)
            stop  = pd.to_timedelta('00:' + stop_str)

            # Duration in seconds
            dur = (stop - start).dt.total_seconds()

            # ---------------- Sequential Start Time Per Half ----------------

            BASE = 19 * 60 + 59  # 19:59

            start_map = {}

            for half, grp in df.groupby('Half'):
                remaining = BASE
                for idx, sec in zip(grp.index, dur.loc[grp.index]):
                    start_map[idx] = remaining
                    remaining = max(0, remaining - sec)

            # Final formatted Time
            df['Time'] = df.index.map(lambda i: f"{int(start_map[i]//60):02}:{int(start_map[i]%60):02}")

            # Drop original Start/Stop columns
            df.drop(columns=['Start', 'Stop'], inplace=True)


            # ---------------- Tie Break Raids ----------------

            tie_cols = ['Yes', 'No']
            for col in tie_cols:
                df[col] = df[col].astype(str).str.strip()  # Remove extra spaces
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

            # 2. Map 1 → skill name, 0 → blank
            for col in tie_cols:
                df[col] = df[col].map({1: col, 0: ''})

            # 3. Join all non-empty skills into a single string
            df['Tie_Break_Raids'] = df[tie_cols].apply(lambda x: ', '.join(filter(None, x)), axis=1)

            df.drop(columns=tie_cols, inplace=True)


            # ---------------- New Columns ----------------
            new_columns = [
                
                # --- Extra Columns ---
                'Video_Link', 'Event', 'Court_Entry',                                 # 3

                # --- TEAM RAID NUMBERING ---
                'Team_Raid_Number', 'Defender_1', 'Defender_2',
                'Defender_3', 'Defender_4', 'Defender_5',
                'Defender_6', 'Defender_7',                                           # 8

                # --- TEAMS & PLAYERS IDENTIFICATION ---
                'Raiding_Team_ID', 'Raiding_Team_Name',
                'Defending_Team_ID', 'Defending_Team_Name', 'Raider_ID',              # 5

                # --- POINTS BREAKDOWN ---
                'Raiding_Team_Points_Pre', 'Defending_Team_Points_Pre',
                'Raiding_Touch_Points', 'Raiding_Bonus_Points',
                'Raiding_Self_Out_Points', 'Raiding_All_Out_Points',
                'Defending_Capture_Points', 'Defending_Bonus_Points',
                'Defending_Self_Out_Points', 'Defending_All_Out_Points',              # 10

                # --- RAID ACTION DETAILS ---
                'Number_of_Raiders', 'Raider_Self_Out',
                'Defenders_Touched_or_Caught'                                         # 3
            ]

            # Add empty new columns
            for col in new_columns:
                df[col] = None
    

            # ---------------- New Logical Order ----------------

            new_order = [
                # 1. Raid Details & Identification
                "Tournament_ID", "Season_ID", "Match_No", "Match_ID", "Event_Number",
                "Match_Raid_Number", "Team_Raid_Number",
            
                # 2. Pre-Raid Points & Time / Half
                "Raiding_Team_Points_Pre", "Defending_Team_Points_Pre", "Half", "Time",
            
                # 3. Team & Raider Info
                "Raiding_Team_ID", "Raiding_Team_Name", "Raider_ID", "Raider_Name",
                "Court_Entry", "Number_of_Raiders", "Number_of_Defenders",
            
                # 4. Raid Details
                "Raid_Number", "Raid_Length", "Outcome", "Bonus", "Type_of_Bonus",
            
                # 5. Raiding Team Points
                "Raiding_Team_Points", "Raiding_Touch_Points", "Raiding_Bonus_Points",
                "Raiding_Self_Out_Points", "Raiding_All_Out_Points",
            
                # 6. Defending Team Points
                "Defending_Team_Points", "Defending_Capture_Points", "Defending_Bonus_Points",
                "Defending_Self_Out_Points", "Defending_All_Out_Points",
            
                # 7. Zone & Team Info
                "Zone_of_Action", "Defending_Team_ID", "Defending_Team_Name",
            
                # 8. Defenders Info
                "Defenders_Touched_or_Caught", "Defender_Position",
                "Defender_1", "Defender_1_Name", "Defender_2", "Defender_2_Name",
                "Defender_3", "Defender_3_Name", "Defender_4", "Defender_4_Name",
                "Defender_5", "Defender_5_Name", "Defender_6", "Defender_6_Name",
                "Defender_7", "Defender_7_Name", "Number_of_Defenders_Self_Out",
            
                # 9. Raider & Skills
                "Raider_Self_Out", "Attacking_Skill", "Defensive_Skill",
                "QoD_Skill", "Counter_Action_Skill",
            
                # 10. Metadata
                "Event", "Technical_Point_Raiding_Team", "Technical_Point_Defending_Team",
                "All_Out", "Tie_Break_Raids", "Video_Link",
            ]

            # Apply the new column order
            df = df[new_order]

  
            # ---------------- Updating Points Columns ----------------

            # Raiding_Bonus_Points
            df["Raiding_Bonus_Points"] = (df["Bonus"] == "Yes").astype(int)

            # Raiding_Touch_Points
            defender_cols = [f'Defender_{i}_Name' for i in range(1, 8)]

            df['Raiding_Touch_Points'] = ((df[defender_cols].notna().sum(axis=1) - df['Number_of_Defenders_Self_Out'])
                .where(df['Outcome'] == 'Successful', 0))
            
            # Convert 'All_Out' column to numeric directly
            df['All_Out'] = pd.to_numeric(df['All_Out'], errors='coerce')
            
            # Update Raiding_All_Out_Points
            df["Raiding_All_Out_Points"] = (((df['Outcome'] == 'Successful') & (df["All_Out"] == 1)).astype(int) * 2)
            
            # Raiding_Self_Out_Points
            df['Raiding_Self_Out_Points'] = df['Number_of_Defenders_Self_Out']

            # Defending_Bonus_Points
            df['Defending_Bonus_Points'] = (((df['Number_of_Defenders'] <= 3) & (df['Outcome'] == 'Unsuccessful')).astype(int))

            # Raider_Self_Out (helper col for defense logic)
            df["Raider_Self_Out"] = (df["Defensive_Skill"] == "Raider self out").astype(int)

            # Defending_Capture_Points
            df['Defending_Capture_Points'] = (((df['Outcome'] == 'Unsuccessful') & (df['Raider_Self_Out'] == 0)).astype(int))

            # Defending_All_Out_Points
            df["Defending_All_Out_Points"] = (((df['Outcome'] == 'Unsuccessful') & (df["All_Out"] == 1)).astype(int) * 2)

            # Defending_Self_Out_Points
            df['Defending_Self_Out_Points'] = df["Raider_Self_Out"]

            # Copy Outcome to Event
            df['Event'] = df['Outcome']

            # Convert Technical Points to numeric
            tech_cols = ["Technical_Point_Raiding_Team", "Technical_Point_Defending_Team"]
            df[tech_cols] = df[tech_cols].apply(pd.to_numeric, errors="coerce").astype("Int64")


            ######## Quality Check #########

            # ----------------------------------------------
            #  Helper Utilities
            # ----------------------------------------------

            def _is_empty(value) -> bool:
                """Return True if a value is NaN, None, or a blank/whitespace string."""
                return pd.isna(value) or str(value).strip() == ""


            def _non_empty_cols(row, columns: list[str]) -> list[str]:
                """Return column names from *columns* that have a non-empty value in *row*."""
                return [c for c in columns if not _is_empty(getattr(row, c))]


            def _col_is_empty(series: pd.Series) -> pd.Series:
                """Vectorized: True where value is NaN, None, or blank/whitespace string."""
                return series.isna() | series.astype(str).str.strip().eq("")


            def _col_is_not_empty(series: pd.Series) -> pd.Series:
                """Vectorized: True where value is NOT empty."""
                return ~_col_is_empty(series)


            def _all_cols_empty(df_sub: pd.DataFrame, columns: list[str]) -> pd.Series:
                """True for rows where ALL listed columns are empty."""
                return pd.DataFrame({c: _col_is_empty(df_sub[c]) for c in columns}).all(axis=1)


            def _any_col_not_empty(df_sub: pd.DataFrame, columns: list[str]) -> pd.Series:
                """True for rows where AT LEAST ONE listed column is not empty."""
                return ~_all_cols_empty(df_sub, columns)


            def _extract_raid_number(value) -> int | None:
                """Extract the integer from strings like 'Raid 5'. Returns None if invalid."""
                if not isinstance(value, str):
                    return None
                match = re.match(r"Raid\s+(\d+)", value.strip())
                return int(match.group(1)) if match else None


            # ----------------------------------------------
            #  Individual QC Checks
            # ----------------------------------------------

            def qc_01_event_sequence(df) -> None:
                """QC 1: Verify raid event numbers form a consecutive sequence."""
                raid_numbers = df["Event_Number"].apply(_extract_raid_number).tolist()
                errors_found = False

                for i in range(1, len(raid_numbers)):
                    prev, curr = raid_numbers[i - 1], raid_numbers[i]
                    label = df.iloc[i]["Event_Number"]

                    if prev is not None and (curr is None or curr != prev + 1):
                        print(f"❌ {label}: Check RAW CSV and Update.\n")
                        errors_found = True

                if not errors_found:
                    print("QC 1: ✅ All rows are Valid.\n")


            def qc_02_empty_columns(df) -> None:
                """QC 2: Key columns must not be Empty."""
                required_cols = [
                    "Raid_Length", "Outcome", "Bonus", "All_Out", "Half","Raid_Number", "Raider_Name", "Number_of_Defenders",
                    "Technical_Point_Raiding_Team", "Technical_Point_Defending_Team", "Tie_Break_Raids"]
                
                empty_mask = pd.DataFrame({c: _col_is_empty(df[c]) for c in required_cols})
                invalid = empty_mask.any(axis=1)

                if invalid.any():
                    for idx in df.index[invalid]:
                        bad = empty_mask.loc[idx][empty_mask.loc[idx]].index.tolist()
                        print(f"\n❌ {df.at[idx, 'Event_Number']}: Empty in → {', '.join(bad)}")
                else:
                    print("\nQC 2: ✅ All rows are Valid.\n")


            def qc_03_empty_outcome_constraints(df) -> None:
                """QC 3: When Outcome = 'Empty', related columns must be Empty."""
                must_be_empty = [
                    "Defender_1_Name", "Defender_2_Name", "Defender_3_Name",
                    "Defender_4_Name", "Defender_5_Name", "Defender_6_Name",
                    "Defender_7_Name", "Zone_of_Action", "Attacking_Skill",
                    "Defensive_Skill", "Counter_Action_Skill","Defender_Position", "QoD_Skill"]
                
                is_outcome_empty = df["Outcome"] == "Empty"
                all_blank = _all_cols_empty(df, must_be_empty)
                invalid = is_outcome_empty & ~(
                    all_blank
                    & (df["All_Out"] == 0)
                    & (df["Raiding_Team_Points"] == 0)
                    & (df["Defending_Team_Points"] == 0)
                    & (df["Bonus"] == "No")
                )

                if invalid.any():
                    for row in df.loc[invalid].itertuples(index=False):
                        issues = []
                        non_empty = _non_empty_cols(row, must_be_empty)
                        if non_empty:
                            issues.append(f" these columns should be empty: {', '.join(non_empty)}")
                        if row.All_Out != 0:
                            issues.append(f"All_Out should be 0 (is {row.All_Out})")
                        if row.Raiding_Team_Points != 0:
                            issues.append(f"Raiding_Team_Points should be 0 (is {row.Raiding_Team_Points})")
                        if row.Defending_Team_Points != 0:
                            issues.append(f"Defending_Team_Points should be 0 (is {row.Defending_Team_Points})")
                        if row.Bonus != "No":
                            issues.append(f"Bonus should be 'No' (is '{row.Bonus}')")
                        print(f"❌ {row.Event_Number}: → When Outcome is 'Empty', ⟶ {' ; '.join(issues)}.\n")
                else:
                    print("QC 3: ✅ All rows are Valid.\n")


            def qc_04_missing_required_fields(df) -> None:
                """QC 4: Successful/Unsuccessful, Bonus=No & Raider Self Out=0 must have following fields."""
                check_cols = ["Defender_1_Name", "Number_of_Defenders", "Zone_of_Action"]
                context = (
                    df["Outcome"].isin(["Successful", "Unsuccessful"])
                    & (df["Bonus"] == "No")
                    & (df["Raider_Self_Out"] == 0)
                )
                some_missing = pd.DataFrame({c: _col_is_empty(df[c]) for c in check_cols}).any(axis=1)
                invalid = context & some_missing

                if invalid.any():
                    for idx, row in df.loc[invalid].iterrows():
                        missing = [c for c in check_cols if _is_empty(row[c])]

                        print(f"❌ {row['Event_Number']}: When Outcome='{row['Outcome']}', No Bonus, No 'Raider Self Out' ⟶ Missing: {', '.join(missing)}.\n")
                else:
                    print("QC 4: ✅ All rows are Valid.\n")


            def qc_05_raid3_requires_empty_two_before(df) -> None:
                """QC 5: If Raid_Number == 3, the row at index −2 must have Outcome == 'Empty'."""
                errors_found = False
                for idx in range(2, len(df)):
                    if df.at[idx, "Raid_Number"] == 3 and df.at[idx - 2, "Outcome"] != "Empty":

                        print(f"❌ {df.at[idx - 2, 'Event_Number']}: → Outcome must be 'Empty' (Because {df.at[idx, 'Event_Number']} has Raid Number = 3)\n")
                        errors_found = True

                if not errors_found:
                    print("QC 5: ✅ All rows are Valid.\n")


            def qc_06_raid1_empty_needs_raid2(df) -> None:
                """QC 6: Raid_Number 1 with Outcome 'Empty' → row at +2 must be Raid_Number 2."""
                errors_found = False
                for idx in range(len(df)):
                    if df.at[idx, "Raid_Number"] == 1 and df.at[idx, "Outcome"] == "Empty":
                        if idx + 2 < len(df) and df.at[idx + 2, "Raid_Number"] != 2:

                            print(f"❌ {df.at[idx + 2, 'Event_Number']}: → Raid Number must be = 2 (Because {df.at[idx, 'Event_Number']} is Empty & Raid Number = 1)\n")
                            errors_found = True

                if not errors_found:
                    print("QC 6: ✅ All rows are Valid.\n")


            def qc_07_success_or_fail_needs_raid1_at_plus2(df) -> None:
                """QC 7: After Successful/Unsuccessful, the row at +2 must have Raid_Number = 1."""
                outcome_clean = df["Outcome"].str.strip().str.lower()
                errors_found = False

                for i in range(len(df) - 2):
                    if outcome_clean.iat[i] in {"successful", "unsuccessful"}:
                        if df.at[i + 2, "Raid_Number"] != 1:

                            print(f"❌ {df.at[i + 2, 'Event_Number']}: Raid Number must be 1 (because {df.at[i, 'Event_Number']} has Outcome='{df.at[i, 'Outcome']}')\n")
                            errors_found = True

                if not errors_found:
                    print("QC 7: ✅ All rows are Valid.\n")


            def qc_08_raid2_empty_requires_raid1_empty(df) -> None:
                """QC 8: Raid_Number 2 & Empty → row at −2 must also be Raid_Number 1 & Empty."""
                outcome_clean = df["Outcome"].str.strip().str.lower()
                errors_found = False

                for i in range(2, len(df)):
                    if df.at[i, "Raid_Number"] == 2 and outcome_clean.iat[i] == "empty":
                        prev_rn = df.at[i - 2, "Raid_Number"]
                        prev_outcome = outcome_clean.iat[i - 2]
                        if not (prev_rn == 1 and prev_outcome == "empty"):

                            print(f"❌ {df.at[i, 'Event_Number']} is Empty, but {df.at[i - 2, 'Event_Number']} has Raid Number={prev_rn} and Outcome='{df.at[i - 2, 'Outcome']}'\n")
                            errors_found = True

                if not errors_found:
                    print("QC 8: ✅ All rows are Valid.\n")


            def qc_09_points_match(df) -> None:
                """QC 9: Sum of Team Point column must equal to Sub Points columns."""

                def _check(cols: list[str], total_col: str, label: str) -> None:
                    calculated = df[cols].sum(axis=1)
                    mismatch = calculated != df[total_col]
                    if mismatch.any():
                        for idx, row in df[mismatch].iterrows():

                            print(f"❌ {row['Event_Number']}: → {label} mismatch (Expected: {df.loc[idx, cols].sum()}, Found: {row[total_col]})\n")
                    else:
                        print(f"QC 9: ✅ All rows are Valid for {label}\n")

                _check(
                    ["Raiding_Touch_Points", "Raiding_Bonus_Points", "Raiding_Self_Out_Points", "Raiding_All_Out_Points", "Technical_Point_Raiding_Team"],
                    "Raiding_Team_Points",
                    "Attacking Points")
                _check(
                    ["Defending_Capture_Points", "Defending_Bonus_Points", "Defending_Self_Out_Points", "Defending_All_Out_Points", "Technical_Point_Defending_Team"],
                    "Defending_Team_Points",
                    "Defensive Points")


            def qc_10_outcome_needs_points(df) -> None:
                """QC 10: Successful/Unsuccessful must have at least one point."""

                def _check(outcome: str, cols: list[str], team: str) -> None:
                    has_outcome = df["Outcome"].eq(outcome)
                    zero_pts = df[cols].fillna(0).sum(axis=1).eq(0)
                    bad = has_outcome & zero_pts
                    if bad.any():
                        for raid_no in df.loc[bad, "Event_Number"].astype(str):
                            print(f"❌ {team}: Raid {raid_no} — Outcome is '{outcome}', but no points were given.\n")
                    else:
                        print(f"QC 10: ✅ All {team} ({outcome}) rows are Valid.\n")

                _check("Successful",
                      ["Raiding_Touch_Points", "Raiding_Bonus_Points", "Raiding_Self_Out_Points", "Raiding_All_Out_Points"],
                      "Raiding")

                _check("Unsuccessful",
                      ["Defending_Capture_Points", "Defending_Bonus_Points", "Defending_Self_Out_Points", "Defending_All_Out_Points"],
                      "Defending")


            def qc_11_defending_points_limit(df) -> None:
                """QC 11: Defending_Self_Out_Points and Defending_Capture_Points must not exceed 1."""
                errors_found = False

                bad_self_out = df["Defending_Self_Out_Points"] > 1
                if bad_self_out.any():
                    for msg in "❌ " + df.loc[bad_self_out, "Event_Number"].astype(str) + "  Check 'Raider self out'\n":
                        print(msg)
                    errors_found = True

                bad_capture = df["Defending_Capture_Points"] > 1
                if bad_capture.any():
                    for msg in "❌ " + df.loc[bad_capture, "Event_Number"].astype(str) + "  Check 'Defensive Points'\n":
                        print(msg)
                    errors_found = True

                if not errors_found:
                    print("QC 11: ✅ All rows are Valid.\n")


            def qc_12_raid_length(df) -> None:
                """QC 12: Raid_Length should be > 2."""
                bad = df["Raid_Length"] <= 2
                if bad.any():
                    for idx in df.index[bad]:
                        print(f"⚠️ {df.at[idx, 'Event_Number']}: 'Raid Length' is {df.at[idx, 'Raid_Length']}\n")
                else:
                    print("QC 12: ✅ All rows have valid Raid Length values.\n")


            def qc_13_defenders_positive(df) -> None:
                """QC 13: Number_of_Defenders should be > 0."""
                bad = df["Number_of_Defenders"] <= 0
                if bad.any():
                    for idx in df.index[bad]:

                        print(f"❌ {df.at[idx, 'Event_Number']}: 'Number of Defenders' is --> {df.at[idx, 'Number_of_Defenders']}, Check \n")
                else:
                    print("QC 13: ✅ All rows are Valid.\n")


            def qc_14_skill_consistency(df) -> None:
                """QC 14: Skill validation with 3 Skill columns."""
            
                subset = df[(df["Outcome"] == "Successful") &
                            (df["Bonus"] == "No") &
                            (df["Number_of_Defenders_Self_Out"] == 0)
                ].copy()
            
                atk_na = _col_is_empty(subset["Attacking_Skill"])
                def_na = _col_is_empty(subset["Defensive_Skill"])
                ca_na = _col_is_empty(subset["Counter_Action_Skill"])
            
                # Case 1: All empty
                all_empty = atk_na & def_na & ca_na
                empty_rows = subset.loc[all_empty, "Event_Number"]
            
                if empty_rows.empty:
                    print("QC 14: ✅ All rows are Valid.\n")
                    return
            
                for event in empty_rows:
                    print(f"❌ {event}: No Skills are Tagged, Check.\n")


            def qc_15_unsuccessful_needs_defensive_skill(df) -> None:
                """QC 15: Unsuccessful outcome must have a non-empty Defensive_Skill."""

                bad = df[(df["Outcome"] == "Unsuccessful") & _col_is_empty(df["Defensive_Skill"])]
                if not bad.empty:
                    for _, row in bad.iterrows():
                        print(f"❌ {row['Event_Number']}: Outcome is 'Unsuccessful' and 'Defensive Skill' is Empty.\n")
                else:
                    print("QC 15: ✅ All rows are Valid.\n")


            def qc_16_defensive_counter_symmetry(df) -> None:
                """QC 16: Successful + No Bonus + Raiding_Touch > 0 → Defensive & Counter skills must both be present or both empty."""

                mask = ((df["Outcome"] == "Successful") & (df["Bonus"] == "No") & (df["Raiding_Touch_Points"] > 0))

                filtered = df[mask]
                def_empty = filtered["Defensive_Skill"].apply(_is_empty)
                ca_empty = filtered["Counter_Action_Skill"].apply(_is_empty)
                violations = filtered[def_empty ^ ca_empty]    # exclusive OR (XOR)

                if violations.empty:
                    print("QC 16: ✅ All rows are Valid.\n")
                else:
                    for event in violations["Event_Number"]:
                        print(f"❌ {event}: 'Defensive Skill' or 'Counter Action Skill' missing.\n")


            def qc_17_defender_position_alignment(df) -> None:
                """QC 17: Defender present ↔ Defender_Position present (both ways)."""
                has_defender = _col_is_not_empty(df["Defender_1_Name"])
                has_position = _col_is_not_empty(df["Defender_Position"])

                fail_no_pos = df[has_defender & ~has_position]
                fail_no_def = df[~has_defender & has_position]

                if fail_no_pos.empty and fail_no_def.empty:
                    print("QC 17: ✅ All 'Defender Positions' are consistent.\n")
                else:
                    for event in fail_no_pos["Event_Number"]:
                        print(f"❌ {event}: Defender(s) present but 'Defender Position' is empty.\n")
                    for event in fail_no_def["Event_Number"]:
                        print(f"❌ {event}: 'Defender Position' present but Defender(s) is empty.\n")


            def qc_18_defensive_qod_alignment(df) -> None:
                """QC 18: When Outcome = Unsuccessful, Defensive_Skill ↔ QoD_Skill must be aligned."""
                excluded = {"Defender self out", "Raider self out"}
                is_unsuccessful = df["Outcome"] == "Unsuccessful"
                has_def = _col_is_not_empty(df["Defensive_Skill"])
                has_qod = _col_is_not_empty(df["QoD_Skill"])
                not_excluded = ~df["Defensive_Skill"].isin(excluded)

                type1 = df[is_unsuccessful & has_def & not_excluded & ~has_qod]
                type2 = df[is_unsuccessful & has_qod & ~has_def]

                if type1.empty and type2.empty:
                    print("QC 18: ✅ 'Defensive Skill' and 'QoD Skill' are aligned correctly.\n")
                else:
                    if not type1.empty:
                        print(f"❌ {type1['Event_Number'].tolist()} → 'Defensive Skill' present but 'QoD Skill' missing.\n")
                    if not type2.empty:
                        print(f"❌ {type2['Event_Number'].tolist()} → 'QoD Skill' present but 'Defensive Skill' missing.\n")


            def qc_19_bonus_type_consistency(df) -> None:
                """QC 19: Bonus = 'Yes' requires non-empty Type_of_Bonus, and vice versa."""
                bonus = df["Bonus"].astype(str).str.strip().str.title()

                yes_missing = (bonus == "Yes") & _col_is_empty(df["Type_of_Bonus"])
                no_present = (bonus == "No") & _col_is_not_empty(df["Type_of_Bonus"])

                failed = df[yes_missing | no_present]

                if failed.empty:
                    print("QC 19: ✅ All rows are Valid.\n")
                else:
                    for _, row in failed.iterrows():
                        b = str(row["Bonus"]).strip().title()
                        t = row["Type_of_Bonus"]
                        if b == "Yes" and _is_empty(t):
                            print(f"❌ {row['Event_Number']}: Bonus is 'Yes' but 'Type of Bonus' is empty.\n")
                        elif b == "No" and not _is_empty(t):
                            print(f"❌ {row['Event_Number']}: Bonus is 'No' but 'Type of Bonus' should be empty.\n")


            def qc_20_zone_required_for_outcome(df) -> None:
                """QC 20: Successful / Unsuccessful must have a non-empty Zone_of_Action."""
                has_outcome = df["Outcome"].isin(["Successful", "Unsuccessful"])
                zone_empty = _col_is_empty(df["Zone_of_Action"])
                bad = df[has_outcome & zone_empty]

                if not bad.empty:
                    for _, row in bad.iterrows():
                        print(f"❌ {row['Event_Number']}: →  'Zone of Action' is Empty.\n")
                else:
                    print(" QC 20: ✅ All rows are Valid.\n")


            def qc_21_raider_self_out_columns_empty(df) -> None:
                """QC 21: When Defensive_Skill = 'Raider self out' and no defenders self-out, specific columns must be empty."""
                cols_to_check = ["QoD_Skill", "Defender_1_Name", "Defender_Position", "Counter_Action_Skill"]

                # Conditions: 
                is_raider_self_out = df["Defensive_Skill"].eq("Raider self out")
                no_defender_self_out = df["Number_of_Defenders_Self_Out"].fillna(0).eq(0) # NaN

                # Combine conditions
                check_mask = is_raider_self_out & no_defender_self_out
                violations = check_mask & _any_col_not_empty(df, cols_to_check)
                flagged = df[violations]

                if not flagged.empty:
                    for _, row in flagged.iterrows():
                        bad_cols = _non_empty_cols(row, cols_to_check)
                        print(f"❌ {row['Event_Number']}: Found values in {', '.join(bad_cols)} — must be Empty.\n")
                else:
                    print("QC 21: ✅ All rows are Valid.\n")


            def qc_22_bonus_only_skills_empty(df) -> None:
                """QC 22: Successful + Bonus Yes + 1 point → all skill columns must be empty."""
                skill_cols = ["Attacking_Skill", "Defensive_Skill", "QoD_Skill", "Counter_Action_Skill"]
                filtered = df[
                    (df["Outcome"] == "Successful")
                    & (df["Bonus"] == "Yes")
                    & (df["Raiding_Team_Points"] == 1)]
                    
                issues_found = False

                for _, row in filtered.iterrows():
                    for col in skill_cols:
                        if not _is_empty(row[col]):
                            print(f"❌ {row['Event_Number']}: When Outcome = 'Successful', Bonus = 'Yes', Attacking Points = 1, all Skill columns must be Empty. But '{col}' has value '{row[col]}'.\n")
                            issues_found = True

                if not issues_found:
                    print("QC 22: ✅ All rows are Valid.\n")


            def qc_23_qod_outcome_alignment(df) -> None:
                """QC 23: QoD_Skill & Outcome alignment — QoD without Defensive, or QoD on Successful."""
                has_def = _col_is_not_empty(df["Defensive_Skill"])
                has_qod = _col_is_not_empty(df["QoD_Skill"])
                is_success = df["Outcome"] == "Successful"

                type1 = df[has_qod & ~has_def]
                type2 = df[is_success & has_qod]

                if type1.empty and type2.empty:
                    print("QC 23: ✅ All Skill and Outcome alignments are Valid.\n")
                else:
                    if not type1.empty:
                        print(f"❌ {type1['Event_Number'].tolist()} → 'QoD Skill' present but 'Defensive Skill' missing.\n")
                    if not type2.empty:
                        print(f"❌ {type2['Event_Number'].tolist()} → Raid is Successful but 'QoD Skill' must be Empty.\n")


            def qc_24_half_sequence(df) -> None:
                """QC 24: Half column must start with 'First' and never go back from 'Second' to 'First'."""
                errors_found = False
                last_val = "First"

                for i, (val, evt) in enumerate(zip(df["Half"], df["Event_Number"])):
                    if i == 0 and val != "First":
                        print(f"❌ {evt}: Sequence must start with 'First'.\n")
                        errors_found = True
                        
                    elif last_val == "Second" and val == "First":
                        print(f"❌ {evt}: Wrong value in 'Half'.\n")
                        errors_found = True
                        
                    elif val not in ("First", "Second"):
                        print(f"❌ {evt}: Invalid value in 'Half' column.\n")
                        errors_found = True
                    last_val = val

                if not errors_found:
                    print("QC 24: ✅ All rows are Valid.\n")


            def qc_25_defender_self_out_rules(df) -> None:
                """QC 25: Defender self-out skill must align with Number_of_Defenders_Self_Out and touch points."""
                qc_failed = False

                for _, row in df.iterrows():
                    # Rule 1: 'Defender self out' needs Number_of_Defenders_Self_Out > 0
                    if (row["Attacking_Skill"] == "Defender self out" and row["Number_of_Defenders_Self_Out"] == 0 and row["Raiding_Self_Out_Points"] == 0):

                        print(f"❌ {row['Event_Number']}: 'Defender self out' requires 'Number of Defenders Self Out' > 0.\n")
                        qc_failed = True

                    # Rule 2: Touch + Self-out points shouldn't coexist with 'Defender self out'
                    if (row["Raiding_Touch_Points"] >= 1 and row["Raiding_Self_Out_Points"] >= 1 and row["Attacking_Skill"] == "Defender self out"):

                        print(f"❌ {row['Event_Number']}: 'Attacking Skill' Must not be 'Defender Self Out'\n")
                        qc_failed = True

                if not qc_failed:
                    print("QC 25: ✅ All rows are Valid.\n")


            def qc_26_defensive_skill_needs_defender(df) -> None:
                """QC 26: Defensive_Skill (except 'Raider self out') requires a Defender_1_Name."""
                qc_failed = False

                for _, row in df.iterrows():
                    if (not _is_empty(row["Defensive_Skill"]) and
                                 row["Defensive_Skill"] != "Raider self out" and _is_empty(row["Defender_1_Name"])):
                        
                        print(f"❌ {row['Event_Number']}: 'Defensive Skill' present but Defender(s) is missing\n")
                        qc_failed = True

                if not qc_failed:
                    print("QC 26: ✅ All rows are Valid.\n")


            def qc_27_bonus_restriction_by_defender_count(df) -> None:
                """QC 27: When Number_of_Defenders is 1–5, Bonus must be 'No' and Type_of_Bonus empty."""
                qc_failed = False

                for _, row in df.iterrows():
                    if row["Number_of_Defenders"] in range(1, 6):
                        if not (row["Bonus"] == "No" and _is_empty(row["Type_of_Bonus"])):

                            print(f"❌ {row['Event_Number']}: 'Number of Defenders' is {row['Number_of_Defenders']}, so Bonus must be 'No' and 'Type of Bonus' must be Empty\n")
                            qc_failed = True

                if not qc_failed:
                    print("QC 27: ✅ All rows are Valid.\n")
                    

            def qc_28_raider_self_out_check(df) -> None:
                """QC 28: When Raider_Self_Out = 1: QoD_Skill and Counter_Action_Skill must be empty, Attacking_Skill must be 'Defender self out' or empty"""
                errors_found = False
            
                for _, row in df.iterrows():
                    if row['Raider_Self_Out'] != 1:
                        continue
                    qod_invalid = not _is_empty(row['QoD_Skill'])
                    counter_invalid = not _is_empty(row['Counter_Action_Skill'])
                    attacking = row['Attacking_Skill']
                    attacking_invalid = not (_is_empty(attacking) or attacking == 'Defender self out')
            
                    if any([qod_invalid, counter_invalid, attacking_invalid]):
                        print(f"❌ {row['Event_Number']}: When 'Raider Self Out' --> 'QoD Skill' & 'Counter Action Skill' must be Empty, and 'Attacking Skill' must be 'Defender self out' or Empty.\n")
                        errors_found = True
            
                if not errors_found:
                    print("QC 28: ✅ All rows are Valid.\n")


            def qc_29_tie_break_raids_check(df) -> None:
                """QC 29: When Tie_Break_Raids = 'Yes', Number_of_Defenders must be 7 and Raid_Number must be 1."""
            
                mask = ((df["Tie_Break_Raids"] == "Yes") &
                       ((df["Number_of_Defenders"] != 7) | (df["Raid_Number"] != 1)))
                flagged = df[mask]
            
                if flagged.empty:
                    print("QC 29: ✅ All rows are Valid.\n")
                    return
                for _, row in flagged.iterrows():
                    msg = " and ".join(filter(None, [
                        "'Number of Defenders' must be 7" if row["Number_of_Defenders"] != 7 else None,
                        "'Raid Number' must be 1" if row["Raid_Number"] != 1 else None]))
                    print(f"❌ {row['Event_Number']}: {msg}.\n")

            # ---------------------------------------------
            #  Main Runner
            # ---------------------------------------------

            def run_all_quality_checks(df: pd.DataFrame) -> None:

                """Execute all 29 quality checks in order on *df*."""

                qc_01_event_sequence(df)
                qc_02_empty_columns(df)
                qc_03_empty_outcome_constraints(df)
                qc_04_missing_required_fields(df)
                qc_05_raid3_requires_empty_two_before(df)
                qc_06_raid1_empty_needs_raid2(df)
                qc_07_success_or_fail_needs_raid1_at_plus2(df)
                qc_08_raid2_empty_requires_raid1_empty(df)
                qc_09_points_match(df)
                qc_10_outcome_needs_points(df)
                qc_11_defending_points_limit(df)
                qc_12_raid_length(df)
                qc_13_defenders_positive(df)
                qc_14_skill_consistency(df)
                qc_15_unsuccessful_needs_defensive_skill(df)
                qc_16_defensive_counter_symmetry(df)
                qc_17_defender_position_alignment(df)
                qc_18_defensive_qod_alignment(df)
                qc_19_bonus_type_consistency(df)
                qc_20_zone_required_for_outcome(df)
                qc_21_raider_self_out_columns_empty(df)
                qc_22_bonus_only_skills_empty(df)
                qc_23_qod_outcome_alignment(df)
                qc_24_half_sequence(df)
                qc_25_defender_self_out_rules(df)
                qc_26_defensive_skill_needs_defender(df)
                qc_27_bonus_restriction_by_defender_count(df)
                qc_28_raider_self_out_check(df)
                qc_29_tie_break_raids_check(df)

            run_all_quality_checks(df)


            # Event_Number formatting
            df['Event_Number'] = (df['Event_Number'].str.extract(r'(\d+)')[0].astype(int).map(lambda x: f"E{x:03d}"))

            df['Video_Link'] = ("https://d3mptnpzpqe58k.cloudfront.net/Pro Kabaddi League/" 
                + df['Season_ID'].astype(str)
                + "/Match " + df['Match_No'].astype(str)
                + "/events/" + (df['Match_Raid_Number'] - 1).astype(int).astype(str).str.zfill(4)
                + "-Raid " + df['Match_Raid_Number'].astype(int).astype(str).str.zfill(3)
                + ".mp4")
            
            # ──────────────────────────────────────────────
            # Finalize and Prepare Excel-Formatting for Download
            # ──────────────────────────────────────────────

            # Prepare final Excel in memory
            excel_buffer = io.BytesIO()
            df.to_excel(excel_buffer, index=False, engine='openpyxl')
            excel_buffer.seek(0)

            # ---- Apply Excel formatting ----
            wb = load_workbook(excel_buffer)
            ws = wb.active

            # Freeze Header Row
            ws.freeze_panes = "A2"

            # Styles
            center = Alignment(horizontal="center", vertical="center")
            bold = Font(bold=True)

            # Apply styles + auto width
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter

                for cell in col:
                    cell.alignment = center
                    cell.border = Border()

                    if cell.row == 1:
                        cell.font = bold

                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = min(max_len + 5, 50)

            # Add filters
            ws.auto_filter.ref = ws.dimensions

            # Save back to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Reset stdout back to default
            sys.stdout = sys.__stdout__

            # --- Show QC logs in scrollable box ---
            qc_text = log_output.getvalue()

            st.markdown(
                f"""
                <div style="height:400px; overflow-y:scroll; border:2px solid yellow;
                            border-radius:8px; padding:10px; background-color:#1E1E1E; color:white;">
                    <pre>{qc_text}</pre>
                </div>
                """,
                unsafe_allow_html=True)

            st.markdown("")
            st.markdown("")

            # --- Show Total Rows and Columns for PROCESSED file ---
            final_rows, final_cols = df.shape if df is not None else (0, 0)
            st.write(f"**Final Total rows:** `{final_rows}` | **Final Total columns:** `{final_cols}`")

            # Show first 5 rows of final file
            st.subheader("Processed File Preview")
            st.dataframe(df, height=210)

            # CSS to style the download button
            st.markdown(
            """
            <style>
            div.stDownloadButton>button {
                color: yellow !important;
                font-weight: bolder !important;
                font-size: 30px !important;  /* Increase font size */
                background-color: black !important;
                border: none !important;
                padding: 10px 20px !important; /* Makes button bigger */
            }
            </style>
            """,
            unsafe_allow_html=True)
            
            st.write(f"**File Name:** `tagged_file_{int(match_no)}.xlsx`")

            # Download button
            st.download_button(
                label="Download Processed Excel",
                data=excel_buffer,
                file_name=f"tagged_file_{int(match_no)}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

        except Exception as e:
            sys.stdout = sys.__stdout__
            st.error(f"❌ An error occurred: {e}")
            
